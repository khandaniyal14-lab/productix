import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def create_factory_mock_excel():
    wb = Workbook()

    # -----------------------------
    # 1️⃣ Products Sheet
    # -----------------------------
    ws1 = wb.active
    ws1.title = "Products"
    ws1.append(["name", "description", "input_fields", "output_fields"])

    products = [
        ["TextileProduction", "Fabric weaving and dyeing", "cotton_kg,dye_liters,water_liters,electricity_kwh", "fabric_meters,waste_kg"],
        ["BeveragePlant", "Soft drink bottling line", "sugar_kg,water_liters,co2_kg,energy_kwh", "bottles,liters_produced"],
        ["BakeryUnit", "Bread and pastry baking", "flour_kg,yeast_kg,energy_kwh,labour_hours", "bread_loaves,waste_kg"]
    ]
    for row in products:
        ws1.append(row)

    for cell in ws1[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # -----------------------------
    # 2️⃣ Batches Sheet
    # -----------------------------
    ws2 = wb.create_sheet("Batches")
    ws2.append(["product_name", "batch_number", "start_date", "end_date", "status"])
    ws2.append(["TextileProduction", "T-1001", "2025-10-10", "2025-10-12", "open"])
    ws2.append(["BeveragePlant", "B-1001", "2025-10-09", "2025-10-10", "open"])
    ws2.append(["BakeryUnit", "BK-1001", "2025-10-07", "2025-10-08", "closed"])

    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # -----------------------------
    # 3️⃣ ShiftEntries Sheet
    # -----------------------------
    ws3 = wb.create_sheet("ShiftEntries")
    ws3.append(["batch_number", "date", "shift_no", "input_materials", "output_products", "admin_notes"])

    ws3.append([
        "T-1001",
        "2025-10-10",
        "Morning",
        '{"cotton_kg": 500, "dye_liters": 25, "water_liters": 2000, "electricity_kwh": 150}',
        '{"fabric_meters": 1200, "waste_kg": 10}',
        "Dyeing line running smoothly"
    ])

    ws3.append([
        "B-1001",
        "2025-10-09",
        "Evening",
        '{"sugar_kg": 100, "water_liters": 800, "co2_kg": 20, "energy_kwh": 90}',
        '{"bottles": 1500, "liters_produced": 1200}',
        "Minor maintenance on CO2 injector"
    ])

    ws3.append([
        "BK-1001",
        "2025-10-07",
        "Night",
        '{"flour_kg": 200, "yeast_kg": 5, "energy_kwh": 60, "labour_hours": 10}',
        '{"bread_loaves": 1800, "waste_kg": 12}',
        "Excellent quality control"
    ])

    for cell in ws3[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # -----------------------------
    # Save file
    # -----------------------------
    file_path = "factory_mock_data.xlsx"
    wb.save(file_path)
    print(f"✅ Excel mock file created successfully: {file_path}")


if __name__ == "__main__":
    create_factory_mock_excel()
