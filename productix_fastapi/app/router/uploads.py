from datetime import date
import io
from fastapi import APIRouter, UploadFile, File, Form, Depends, HTTPException
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.orm import Session
import pandas as pd
import uuid, hashlib, os
from typing import Optional
from ..deps import get_db, get_current_user
from .. import models  # your SQLAlchemy models

import os
import io
import uuid
import json
import hashlib
from datetime import date

import pandas as pd
from fastapi import APIRouter, File, UploadFile, Depends, HTTPException, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from .. import models
 # adjust import if different

router = APIRouter(prefix="/api/v1/uploads", tags=["uploads"])

UPLOAD_DIR = "/tmp/uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)


# ---------------------------------------------------------------------
# Utility
# ---------------------------------------------------------------------
def compute_import_hash(file_id: str, row_index: int) -> str:
    """Generate a hash for each row (optional helper)."""
    return hashlib.sha256(f"{file_id}:{row_index}".encode()).hexdigest()


# ---------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------
def validate_workbook(sheets: dict):
    errors = []
    warnings = []

    required_sheets = ["Products", "Batches", "ShiftEntries"]
    for sheet in required_sheets:
        if sheet not in sheets:
            errors.append(f"Missing required sheet: {sheet}")

    if "Products" in sheets:
        required_cols = {"name", "input_fields", "output_fields"}
        missing = required_cols - set(sheets["Products"].columns)
        if missing:
            errors.append(f"'Products' missing columns: {missing}")

    if "Batches" in sheets:
        required_cols = {"batch_number", "product_name"}
        missing = required_cols - set(sheets["Batches"].columns)
        if missing:
            errors.append(f"'Batches' missing columns: {missing}")

    if "ShiftEntries" in sheets:
        required_cols = {"batch_number", "shift_no", "date"}
        missing = required_cols - set(sheets["ShiftEntries"].columns)
        if missing:
            errors.append(f"'ShiftEntries' missing columns: {missing}")

    return {"errors": errors, "warnings": warnings}


# ---------------------------------------------------------------------
# Processor
# ---------------------------------------------------------------------
import json
from sqlalchemy.exc import SQLAlchemyError
from datetime import datetime

def process_workbook(db: Session, organization_id: int, file_id: str, sheets: dict, replace_mode: str):
    """Process Excel workbook into DB with validation and logging."""

    batches_df = sheets.get("Batches")
    shifts_df = sheets.get("ShiftEntries")

    if batches_df is None or shifts_df is None:
        raise HTTPException(400, "Missing required sheets (Batches or ShiftEntries)")

    # --- Clean NaNs ---
    batches_df = batches_df.fillna("")
    shifts_df = shifts_df.fillna("")

    batches_created = 0
    shifts_created = 0
    errors = []

    print("📘 Starting Excel import...")
    print(f"Total batches: {len(batches_df)}, Total shifts: {len(shifts_df)}")

    for _, batch_row in batches_df.iterrows():
        try:
            product_name = str(batch_row["product_name"]).strip()
            batch_number = str(batch_row["batch_number"]).strip()

            if not product_name or not batch_number:
                errors.append(f"Missing product_name or batch_number in batch row: {batch_row.to_dict()}")
                continue

            print(f"➡️ Processing batch: {batch_number} ({product_name})")

            # --- Find or create product ---
            product = db.query(models.Product).filter_by(
                name=product_name,
                organization_id=organization_id
            ).first()

            if not product:
                print(f"🆕 Creating new product: {product_name}")
                product = models.Product(
                    name=product_name,
                    organization_id=organization_id,
                    description=batch_row.get("description", ""),
                    input_fields={},
                    output_fields={}
                )
                db.add(product)
                db.commit()
                db.refresh(product)

            # --- Find or create batch ---
            batch = db.query(models.Batch).filter_by(
                batch_number=batch_number,
                organization_id=organization_id
            ).first()

            if not batch:
                batch = models.Batch(
                    batch_number=batch_number,
                    organization_id=organization_id,
                    product_id=product.id,
                    start_date=batch_row.get("start_date") or datetime.utcnow().date(),
                    end_date=batch_row.get("end_date") or None,
                    status=str(batch_row.get("status", "open")).lower()
                )
                db.add(batch)
                db.commit()
                db.refresh(batch)
                batches_created += 1
                print(f"✅ Created batch: {batch_number}")
            else:
                print(f"↩️ Using existing batch: {batch_number}")

            # --- Process shifts for this batch ---
            batch_shifts = shifts_df[shifts_df["batch_number"] == batch_number]
            if batch_shifts.empty:
                print(f"⚠️ No shift entries found for batch {batch_number}")
                continue

            for _, shift_row in batch_shifts.iterrows():
                try:
                    # Parse JSON safely
                    def safe_json(value):
                        if not value or pd.isna(value): 
                            return {}
                        if isinstance(value, dict):
                            return value
                        if isinstance(value, str):
                            try:
                                return json.loads(value)
                            except json.JSONDecodeError:
                                print(f"⚠️ Invalid JSON: {value}")
                                return {}
                        return {}

                    input_parsed = safe_json(shift_row.get("input_materials"))
                    output_parsed = safe_json(shift_row.get("output_products"))

                    date_str = shift_row.get("date")
                    shift_date = None
                    try:
                        shift_date = pd.to_datetime(date_str).date() if date_str else datetime.utcnow().date()
                    except Exception:
                        shift_date = datetime.utcnow().date()

                    shift_entry = models.ShiftEntry(
                        batch_id=batch.id,
                        date=shift_date,
                        shift_no=str(shift_row.get("shift_no", "Morning")),
                        organization_id=organization_id,
                        input_materials=input_parsed,
                        output_products=output_parsed,
                        admin_notes=str(shift_row.get("admin_notes", ""))
                    )
                    db.add(shift_entry)
                    shifts_created += 1
                except Exception as shift_err:
                    err_msg = f"Shift insert failed for batch {batch_number}: {shift_err}"
                    print("❌", err_msg)
                    errors.append(err_msg)
                    db.rollback()

            db.commit()

        except SQLAlchemyError as db_err:
            db.rollback()
            err_msg = f"DB error for batch {batch_row.get('batch_number')}: {db_err}"
            print("❌", err_msg)
            errors.append(err_msg)
        except Exception as e:
            db.rollback()
            err_msg = f"Unexpected error: {e}"
            print("❌", err_msg)
            errors.append(err_msg)

    print(f"✅ Import done: {batches_created} batches, {shifts_created} shifts created")

    return {
        "status": "completed_with_warnings" if errors else "completed",
        "batches_created": batches_created,
        "shifts_created": shifts_created,
        "errors": errors
    }



# ---------------------------------------------------------------------
# Excel Upload Endpoint
# ---------------------------------------------------------------------




@router.post("/excel")
async def upload_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user = Depends(get_current_user)
):
    try:
        # --- Step 1: Load Excel file ---
        df_dict = pd.read_excel(file.file, sheet_name=None)
        expected_sheets = ["Products", "Batches", "ShiftEntries"]
        found_sheets = list(df_dict.keys())
        print(f"✅ SHEETS FOUND: {found_sheets}")

        for sheet in expected_sheets:
            if sheet not in df_dict:
                return JSONResponse(
                    status_code=400,
                    content={"error": f"Missing required sheet: {sheet}"}
                )

        # --- Step 2: Upload Products ---
        products_df = df_dict["Products"].fillna("")
        print(f"📦 Products sheet rows: {len(products_df)}")

        # Build a product map: normalized name -> product
        product_map = {}
        existing_products = db.query(models.Product).filter_by(
            organization_id=user.organization_id
        ).all()
        for p in existing_products:
            product_map[p.name.strip().lower()] = p

        for _, row in products_df.iterrows():
            name_norm = row["name"].strip().lower()
            product = product_map.get(name_norm)

            input_fields = []
            output_fields = []
            try:
                if row.get("input_fields"):
                    input_fields = json.loads(row["input_fields"]) if isinstance(row["input_fields"], str) else row["input_fields"]
                if row.get("output_fields"):
                    output_fields = json.loads(row["output_fields"]) if isinstance(row["output_fields"], str) else row["output_fields"]
            except Exception:
                pass

            if product:
                # Update existing product
                product.description = row.get("description", product.description)
                product.input_fields = input_fields
                product.output_fields = output_fields
            else:
                # Create new product
                product = models.Product(
                    name=row["name"].strip(),
                    description=row.get("description", ""),
                    input_fields=input_fields,
                    output_fields=output_fields,
                    organization_id=user.organization_id
                )
                db.add(product)
                db.flush()  # assign ID without committing
                product_map[name_norm] = product

        db.commit()

        # --- Step 3: Upload Batches ---
        batches_df = df_dict["Batches"].fillna("")
        print(f"📦 Batches sheet rows: {len(batches_df)}")

        for _, row in batches_df.iterrows():
            prod_name_norm = row["product_name"].strip().lower()
            product = product_map.get(prod_name_norm)
            if not product:
                print(f"⚠️ Product not found for batch: {row['batch_number']}")
                continue
            print(f"➡️ Processing batch: {row['batch_number']} for product: {product.name}")
            batch = db.query(models.Batch).filter_by(batch_number=row["batch_number"]).first()
            if batch:
                print(f"↩️ Batch exists: {batch.batch_number}, updating...")
                batch.start_date = row.get("start_date", batch.start_date)
                batch.end_date = row.get("end_date", batch.end_date)
                batch.status = row.get("status", batch.status)
            else:
                batch = models.Batch(
                    organization_id=user.organization_id,
                    product_id=product.id,
                    batch_number=row["batch_number"],
                    start_date=row.get("start_date"),
                    end_date=row.get("end_date"),
                    status=row.get("status", "open")
                )
                print("✅ Creating batch:", batch.batch_number)
                db.add(batch)
                print("✅ Created batch:", batch.batch_number)

        db.commit()

        # --- Step 4: Upload Shift Entries ---
        shifts_df = df_dict["ShiftEntries"].fillna("")
        print(f"📦 ShiftEntries sheet rows: {len(shifts_df)}")

        for _, row in shifts_df.iterrows():
            batch = db.query(models.Batch).filter_by(batch_number=row["batch_number"]).first()
            if not batch:
                print(f"⚠️ Batch not found for shift entry: {row['shift_no']} on {row['date']}")
                continue

            shift = db.query(models.ShiftEntry).filter_by(
                batch_id=batch.id,
                date=row["date"],
                shift_no=row["shift_no"]
            ).first()

            input_materials = {}
            output_products = {}
            try:
                if row.get("input_materials"):
                    input_materials = json.loads(row["input_materials"]) if isinstance(row["input_materials"], str) else row["input_materials"]
                if row.get("output_products"):
                    output_products = json.loads(row["output_products"]) if isinstance(row["output_products"], str) else row["output_products"]
            except Exception:
                pass

            if shift:
                # Update existing shift
                shift.input_materials = input_materials
                shift.output_products = output_products
                shift.admin_notes = row.get("admin_notes", shift.admin_notes)
            else:
                # Create new shift
                shift = models.ShiftEntry(
                    batch_id=batch.id,
                    organization_id=user.organization_id,
                    date=row["date"],
                    shift_no=row["shift_no"],
                    input_materials=input_materials,
                    output_products=output_products,
                    admin_notes=row.get("admin_notes", "")
                )
                db.add(shift)

        db.commit()

        return {"success": True, "message": "Excel uploaded and processed successfully."}

    except Exception as e:
        print("❌ Upload failed:", e)
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------
# Template Download Endpoint
# ---------------------------------------------------------------------
@router.get("/template")
def download_excel_template():
    """
    Download a standard Excel template for uploading product, batch, and shift data.
    Sheets:
    - Products
    - Batches
    - ShiftEntries
    """

    wb = Workbook()

    # 🧩 Products
    ws1 = wb.active
    ws1.title = "Products"
    ws1.append(["name", "description", "input_fields", "output_fields"])

    # Example products with input/output fields as JSON lists
    ws1.append([
        "Widget A",
        "First product",
        json.dumps(["steel", "plastic"]),
        json.dumps(["finished_goods"])
    ])
    ws1.append([
        "Widget B",
        "Second product",
        json.dumps(["cotton", "paint"]),
        json.dumps(["cloth_pack"])
    ])

    for cell in ws1[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # 🧩 Batches
    ws2 = wb.create_sheet("Batches")
    ws2.append(["product_name", "batch_number", "start_date", "end_date", "status"])
    ws2.append(["Widget A", "BATCH-001", date.today().isoformat(), date.today().isoformat(), "open"])
    ws2.append(["Widget B", "BATCH-002", date.today().isoformat(), date.today().isoformat(), "closed"])

    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # 🧩 Shift Entries
    ws3 = wb.create_sheet("ShiftEntries")
    ws3.append([
        "batch_number", "date", "shift_no",
        "input_materials", "output_products", "admin_notes"
    ])
    ws3.append([
        "BATCH-001", date.today().isoformat(), "Morning",
        json.dumps({"steel": {"amount": 100, "unit_price": 5.0}, "plastic": {"amount": 50, "unit_price": 3.0}}),
        json.dumps({"finished_goods": {"amount": 130}}),
        "Smooth run"
    ])
    ws3.append([
        "BATCH-002", date.today().isoformat(), "Evening",
        json.dumps({"cotton": {"amount": 200, "unit_price": 4.0}, "paint": {"amount": 20, "unit_price": 2.0}}),
        json.dumps({"cloth_pack": {"amount": 210}}),
        "Minor issues"
    ])

    for cell in ws3[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Write to stream
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="upload_template.xlsx"'}
    )
   
